"""
EFSA PRIMo Rev 3.1 Excel importer.
Populates primo_commodities and reg396_commodities from the PRIMo workbook.

Download PRIMo Rev 3.1 from:
    https://www.efsa.europa.eu/en/applications/pesticides/tools
    (look for 'Pesticide Residue Intake Model' — downloads as a .xlsx)

Sheet names in PRIMo Rev 3.1 (verify these match your download):
    'Commodity list'   — contains Annex I codes, commodity names, unit weights
    'Processing'       — processing factors (optional, not imported here)

Column layout in 'Commodity list' (approximate — verify headers in your file):
    Col A: Annex I code (e.g. '0110010')
    Col B: Commodity name (official Annex I English name)
    Col C: PRIMo commodity name (may differ slightly from Annex I)
    Col D: Unit weight (g) — used for IESTI calculation
    Col E: Hierarchy level indicator (blank = group/subgroup, filled = individual)
    ... further columns: consumption data by population group (not needed here)

Usage:
    pip install openpyxl
    python import_primo.py --xlsx PRIMo_Rev3.1.xlsx --db crop_db.sqlite

    Add --apply to write to database (default is dry run / preview only).
"""

import argparse, re, sqlite3, sys

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# --- Adjust these if your PRIMo version has different sheet/column layout ---
SHEET_COMMODITY  = "Commodity list"   # Sheet name
COL_ANNEX1_CODE  = 1   # Column index (1-based) for Reg 396 Annex I code
COL_ANNEX1_NAME  = 2   # Column index for official Annex I commodity name
COL_PRIMO_NAME   = 3   # Column index for PRIMo display name
COL_UNIT_WEIGHT  = 4   # Column index for unit weight (g)
HEADER_ROWS      = 3   # Number of header rows to skip


def parse_hierarchy_level(code):
    """
    Infer Annex I hierarchy level from the numeric code pattern.
    Level 1: 7-digit ending in 000 (e.g. 0110000)
    Level 2: 7-digit ending in 0 but not 000 (e.g. 0110010 is actually level 3)
    Actually the EFSA convention is:
      xxxxxxx0 with last 3 zeros = major group (level 1)
      xxxxxxx0 with last 2 zeros = subgroup (level 2, rare)
      xxxxxxx  no trailing zeros = individual commodity (level 3)
    """
    if not code or not re.match(r'^\d{7}$', str(code)):
        return None
    c = str(code)
    if c.endswith('000'):
        return 1
    elif c.endswith('00'):
        return 2
    else:
        return 3


def parent_code(code):
    """Derive parent Annex I code from child code."""
    c = str(code)
    if not re.match(r'^\d{7}$', c) or c.endswith('000'):
        return None
    # Walk up: strip rightmost non-zero digits
    # e.g. 0110010 -> 0110000 (group)
    # This is heuristic; exact parent must be verified against Annex I text.
    if c.endswith('00'):
        return c[:4] + '000'
    elif c[-2:] != '00':
        return c[:5] + '000'
    return None


def import_primo(xlsx_path, db_path, dry_run=True, primo_version="Rev 3.1"):
    if not HAS_OPENPYXL:
        sys.exit("openpyxl not installed. Run: pip install openpyxl")

    print(f"Opening {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    if SHEET_COMMODITY not in wb.sheetnames:
        print(f"Sheet '{SHEET_COMMODITY}' not found. Available sheets:")
        for s in wb.sheetnames:
            print(f"  - {s}")
        sys.exit("Adjust SHEET_COMMODITY constant.")

    ws = wb[SHEET_COMMODITY]

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    cur  = conn.cursor()

    # Get or create a generic 'Unknown' crop as placeholder for unmapped commodities
    cur.execute("SELECT crop_id FROM crops WHERE common_name_en = 'Unknown (unmapped)'")
    row = cur.fetchone()
    if not row:
        cur.execute("""
            INSERT INTO crops (common_name_en, scientific_name, is_food_crop, notes)
            VALUES ('Unknown (unmapped)', NULL, 0,
                    'Placeholder for PRIMo commodities not yet matched to a crop_id. '
                    'Update crop_id after reviewing the commodity name.')
        """)
        unknown_crop_id = cur.lastrowid
    else:
        unknown_crop_id = row[0]

    primo_rows = []
    reg396_rows = []
    preview_count = 0

    for i, row in enumerate(ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True)):
        if not any(row):
            continue

        annex1_code  = str(row[COL_ANNEX1_CODE - 1] or "").strip()
        annex1_name  = str(row[COL_ANNEX1_NAME - 1] or "").strip()
        primo_name   = str(row[COL_PRIMO_NAME  - 1] or "").strip()
        unit_weight  = row[COL_UNIT_WEIGHT - 1]

        if not annex1_code or not annex1_name:
            continue

        # Normalise code — sometimes stored as integer in Excel
        if re.match(r'^\d{1,7}$', annex1_code):
            annex1_code = annex1_code.zfill(7)

        hlevel = parse_hierarchy_level(annex1_code)
        pcode  = parent_code(annex1_code)

        if preview_count < 10:
            print(f"  Preview row {i}: code={annex1_code} name={annex1_name[:40]} "
                  f"primo={primo_name[:30]} L{hlevel}")
            preview_count += 1

        reg396_rows.append({
            "annex1_code":       annex1_code,
            "annex1_name":       annex1_name,
            "hierarchy_level":   hlevel or 3,
            "parent_annex1_code": pcode,
            "crop_id":           unknown_crop_id,
            "regulation_version": "Reg (EC) No 396/2005 (as in PRIMo Rev 3.1)",
        })

        if primo_name:
            primo_rows.append({
                "crop_id":      unknown_crop_id,
                "primo_version": primo_version,
                "primo_code":   annex1_code,
                "primo_name":   primo_name,
                "unit_weight_g": float(unit_weight) if unit_weight else None,
            })

    print(f"\nParsed {len(reg396_rows)} reg396 commodities, {len(primo_rows)} primo entries")

    if dry_run:
        print("(Dry run — pass --apply to write to DB)")
        return

    # Insert reg396_commodities
    inserted_r, skipped_r = 0, 0
    for r in reg396_rows:
        try:
            cur.execute("""
                INSERT OR IGNORE INTO reg396_commodities
                    (crop_id, annex1_code, annex1_name, hierarchy_level,
                     parent_annex1_code, regulation_version)
                VALUES (:crop_id, :annex1_code, :annex1_name, :hierarchy_level,
                        :parent_annex1_code, :regulation_version)
            """, r)
            inserted_r += 1
        except Exception as e:
            print(f"  reg396 skip {r['annex1_code']}: {e}")
            skipped_r += 1

    # Insert primo_commodities
    inserted_p, skipped_p = 0, 0
    for p in primo_rows:
        try:
            cur.execute("""
                INSERT OR IGNORE INTO primo_commodities
                    (crop_id, primo_version, primo_code, primo_name, unit_weight_g)
                VALUES (:crop_id, :primo_version, :primo_code, :primo_name, :unit_weight_g)
            """, p)
            inserted_p += 1
        except Exception as e:
            print(f"  primo skip {p['primo_name'][:30]}: {e}")
            skipped_p += 1

    conn.commit()
    conn.close()

    print(f"\nInserted: {inserted_r} reg396 rows, {inserted_p} primo rows")
    print(f"Skipped:  {skipped_r} reg396 rows, {skipped_p} primo rows")
    print(f"\nNEXT STEP: Update crop_id on reg396_commodities and primo_commodities")
    print("  rows where crop_id = {unknown_crop_id} (the placeholder).")
    print("  Use EPPO names or scientific names to match to your crops table,")
    print("  or run the eppo_verify.py script to cross-reference.")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Import PRIMo Rev 3.1 commodity list into crop DB")
    p.add_argument("--xlsx",    required=True, help="Path to PRIMo Rev 3.1 .xlsx file")
    p.add_argument("--db",      default="crop_db.sqlite")
    p.add_argument("--version", default="Rev 3.1", help="PRIMo version label (default: 'Rev 3.1')")
    p.add_argument("--apply",   action="store_true", help="Write to DB (default: dry run)")
    args = p.parse_args()
    import_primo(args.xlsx, args.db, dry_run=not args.apply, primo_version=args.version)
